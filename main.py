import pandas as pd
from openpyxl import Workbook, load_workbook
import glob,oracledb,os
#INIT LIBRARY
oracledb.init_oracle_client(lib_dir="C:\oracle\instantclient_23_8")
#DATABASE CONNECTION
connection = oracledb.connect(
    user="YOUR_USER",
    password="YOUR",
    dsn="localhost:1521/XE"
)
cursor = connection.cursor()
#CHANGE FORMAT CSV TO XLSX
os.makedirs('Organizador', exist_ok=True)


#TRANSFORM CSV FORMAT TO XLSX ( UNCOMMENT IF YOU HAVE CSV FORMAT  )

# for file in os.listdir("Excels"):
#     if file.endswith(".csv"):
#         enter_route = os.path.join("Excels",file)

#         try:
#             #READ CSV
#             df = pd.read_csv(enter_route)
#             #OUTPUT
#             file_name = os.path.splitext(file)[0] + ".xlsx"
#             exit_route = os.path.join("Organizador",file_name)
#             #SAVE FILE
#             df.to_excel(exit_route,index=False,engine="openpyxl")
#             print("Archivo convertido!")        
#         except Exception as err:
#             print("Ocurrio el siguiente error: ", err)

#MAP FOR PYTHON FORMAT TO SQL FORMAT
def mapping(python_type):
    map = {
        "str":"VARCHAR2(255)",
        "int":"INT",
        "float":"FLOAT",
        "bool":"BOOLEAN"
    }
    return map.get(python_type,"TEXT") #TEXT IF PYTHON DONT RECOGNIZE THE TYPE

#GET THE HEADER AND THE VALUE BELOW
def column_names(sh, data_file= 2):
    colNames = {}
    for col in sh.iter_cols(min_row = 1 , max_row = data_file):
        column_name = col[0].value
        value = col[1].value

        colNames[column_name] = value
    return colNames

#DICT CREATION
def create_dict(data):
    value_dict = {}
    try:
        for i in data:

            value_dict[i] = {
                "valor" : data.get(i),
                "tipo"  : type(data.get(i)).__name__
            }
    except Exception as err:
        print(f"Error en el campo '{i}': {err}")
    return value_dict
    
#EXTRACT THE VALUES
def extract_values(sh):
    values = []
    headers = [cell.value for cell in sh[1]]
    for row in sh.iter_rows(min_row = 2 ,values_only = True):
        if any(row):
            value = dict(zip(headers,row))
            values.append(value)
    return values

#GENERATE THE INSERTS TO SQL
def generate_insert(table_name, values):
    sentences = []
    for fila in values:
        columnas = ", ".join(fila.keys())
        data_values = []

        for q in fila.values():
            if isinstance(q, str):
                popValue = q.replace("'","''")
                data_values.append(f"'{popValue}'")
            elif q is None:
                data_values.append("NULL")
            else:
                data_values.append(str(q))

        sql_values = ", ".join(data_values)
        sentence = f"INSERT INTO {table_name} ({columnas}) VALUES ({sql_values})"
        sentences.append(sentence)

    return sentences

#OPEN FILES
files = glob.glob("./Organizador/*.xlsx")
  
#MAIN
for file in files:
    try:
        sql_camps = []
        new_dict = {}
        wb = load_workbook(file)
        sh = wb.worksheets[0]
        new_sheet_name = file[14:21]
        data = column_names(sh)
        new_dict = create_dict(data)
        for camp , values in new_dict.items():
            value_type = mapping(values['tipo'])
            sql_camps.append(f'{camp} {value_type}')
        table_query = f"CREATE TABLE {new_sheet_name} (\n" + ",\n".join(sql_camps) + "\n)"
        cursor.execute(table_query)
        print(f"\nTabla {new_sheet_name} creada exitosamente!\n")
        register = extract_values(sh)
        inserts = generate_insert(new_sheet_name,register)
        
        for ins in inserts:
            try:
                cursor.execute(ins)
            except Exception as err:
                print(f"Error al insertar Datos: {err}")
        print(f"Datos de la tabla {new_sheet_name} Insertados!")
        connection.commit()

    except Exception as err:
        print(f"Ocurrio un error durante la conversion: {err}")
cursor.close()